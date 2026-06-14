#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None


HEADER_ALIASES = {
    "住宿/": "住宿",
    "住宿费": "住宿",
    "住宿/补贴": "住宿",
}

DEFAULT_HEADERS = [
    "日期",
    "客户",
    "公里数",
    "油费",
    "打车",
    "过路费",
    "餐饮费",
    "礼品费",
    "停车",
    "住宿",
    "补贴",
    "总费用",
    "备注",
]

EXPENSE_COLUMNS = [
    {
        "key": "油费",
        "oa_category": "交通费（含跨地区汽车费）",
        "detail_type": "私车公用-油费",
        "attachments": "高德历史导航截图；成品油发票",
        "risk": "油票金额需大于或等于公里数×1.4，发票日期尽量接近出差日期。",
    },
    {
        "key": "打车",
        "oa_category": "交通费（含跨地区汽车费）",
        "detail_type": "打车软件",
        "attachments": "打车行程单；打车电子发票",
        "risk": "打车软件需行程单和发票成对；金额为“？”时先补金额和票据。",
    },
    {
        "key": "过路费",
        "oa_category": "交通费（含跨地区汽车费）",
        "detail_type": "私车公用-其他相关费用/过路费",
        "attachments": "ETC电子票据汇总单；通行费电子发票",
        "risk": "ETC需附汇总单，通行日期应与自驾拜访日期一致。",
    },
    {
        "key": "餐饮费",
        "oa_category": "交际应酬费",
        "detail_type": "餐费",
        "attachments": "餐饮发票；付款记录（后开票/必要时）；客户与同事名单",
        "risk": "需注明客户公司、客户姓名、同事姓名、午餐/晚餐和业务背景。",
    },
    {
        "key": "礼品费",
        "oa_category": "交际应酬费",
        "detail_type": "礼品费",
        "attachments": "礼品发票；订单/商品明细；客户公司与人员名单",
        "risk": "礼品费需商品明细和事由；无明细容易被退回。",
    },
    {
        "key": "停车",
        "oa_category": "交通费（含跨地区汽车费）",
        "detail_type": "停车费",
        "attachments": "停车票/电子发票；付款记录（必要时）",
        "risk": "停车费单独列明日期、地点和金额，不要混入油费。",
    },
    {
        "key": "住宿",
        "oa_category": "住宿费",
        "detail_type": "住宿费",
        "attachments": "酒店水单+发票；无水单则订单截图/批准",
        "risk": "住宿需水单+发票；超标时写明原因。",
    },
    {
        "key": "补贴",
        "oa_category": "差旅餐饮补贴",
        "detail_type": "业务差旅",
        "attachments": "出差起止时间、出差地和扣除报销餐说明",
        "risk": "出差当天已有客户餐费时需扣除对应餐补。",
    },
]


@dataclass
class FileInfo:
    path: str
    ext: str
    size: int
    categories: list[str]
    metadata: dict[str, Any]


def normalize_header(value: Any, index: int) -> str:
    text = str(value or "").strip()
    if not text:
        return DEFAULT_HEADERS[index - 1] if index <= len(DEFAULT_HEADERS) else f"列{index}"
    return HEADER_ALIASES.get(text, text)


def cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def text_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return str(round(value, 2))
    return str(value).strip()


def number_value(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "").replace("¥", "")
        if cleaned in {"", "？", "?"}:
            return None
        try:
            return round(float(cleaned), 2)
        except ValueError:
            return None
    return None


def classify_file(rel_path: str) -> list[str]:
    categories: list[str] = []
    if re.search(r"话费|电话|手机", rel_path, re.I):
        categories.append("电话费")
    if "停车" in rel_path:
        categories.append("停车费")
    if re.search(r"油|石化|石油|汽油|加油站", rel_path, re.I):
        categories.append("油票")
    if re.search(r"通行|过路|高速|ETC|票根", rel_path, re.I):
        categories.append("过路费")
    if re.search(r"酒店|住宿|结账|水单", rel_path, re.I):
        categories.append("住宿")
    if re.search(r"餐|饭|料理|星巴克|面点王|太二|柴湖|陈鹏鹏|小吃|餐饮", rel_path, re.I):
        categories.append("餐费/交际")
    if re.search(r"滴滴|曹操|阳光|享道|高铁|火车|行程|铁路|客票|12306", rel_path, re.I):
        categories.append("交通")
    if re.search(r"高德|导航|足迹|里程", rel_path, re.I):
        categories.append("里程截图")
    return categories


def parse_pdf_metadata(path: Path) -> dict[str, Any]:
    if path.suffix.lower() != ".pdf" or PdfReader is None:
        return {}
    try:
        reader = PdfReader(str(path))
        text = "\n".join((page.extract_text() or "") for page in reader.pages[:2])
    except Exception:
        return {}

    compact = " ".join(line.strip() for line in text.splitlines() if line.strip())
    metadata: dict[str, Any] = {}

    date_match = re.search(r"20\d{2}年\d{1,2}月\d{1,2}日", compact)
    if date_match:
        metadata["invoice_date"] = date_match.group(0)

    preferred_amount: float | None = None
    total_match = re.search(r"小\s*写\s*[）)]\s*([0-9]+(?:\.[0-9]{1,2})?)\s*[¥￥]", compact)
    if total_match:
        try:
            preferred_amount = float(total_match.group(1))
        except ValueError:
            preferred_amount = None

    amounts: list[float] = []
    for left, right in re.findall(r"[（(]小写[）)]\s*[¥￥]?\s*([0-9]+(?:\.[0-9]{1,2})?)|[¥￥]\s*([0-9]+(?:\.[0-9]{1,2})?)", compact):
        value = left or right
        try:
            amounts.append(float(value))
        except ValueError:
            pass
    ticket_price = re.search(r"票价\s*[:：]\s*[¥￥]\s*([0-9]+(?:\.[0-9]{1,2})?)", compact)
    if ticket_price:
        try:
            amounts.append(float(ticket_price.group(1)))
        except ValueError:
            pass
    trip_total = re.search(r"合计\s*([0-9]+(?:\.[0-9]{1,2})?)元", compact)
    if trip_total:
        try:
            amounts.append(float(trip_total.group(1)))
        except ValueError:
            pass
    if preferred_amount is not None:
        metadata["amount"] = round(preferred_amount, 2)
    elif amounts:
        metadata["amount"] = round(max(amounts), 2)

    invoice_no = re.search(r"发票号码[：:\s]*([0-9]{8,30})", compact)
    if invoice_no:
        metadata["invoice_no"] = invoice_no.group(1)

    if "德凯质量认证（上海）有限公司广州分公司" in compact:
        metadata["buyer"] = "德凯质量认证（上海）有限公司广州分公司"

    seller = re.search(r"91440101687698432H\s*([^0-9¥]{4,80}?)(?:9[0-9A-Z]{14,20}|[0-9A-Z]{15,20})", compact)
    if seller:
        metadata["seller"] = seller.group(1).strip()

    didi_travel_date = re.search(r"出行日期\s+(\d{4}-\d{2}-\d{2})", compact)
    trip_range_date = re.search(r"行程起止日期[：:]\s*(\d{4}-\d{2}-\d{2})", compact)
    railway_date = re.search(r"(20\d{2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", compact)
    if didi_travel_date:
        metadata["travel_date"] = didi_travel_date.group(1)
    elif trip_range_date:
        metadata["travel_date"] = trip_range_date.group(1)
    elif "铁路电子客票" in compact and railway_date:
        metadata["travel_date"] = f"{railway_date.group(1)}-{railway_date.group(2).zfill(2)}-{railway_date.group(3).zfill(2)}"
    elif "didi" in compact.lower() or "滴滴" in compact:
        any_iso_date = re.search(r"20\d{2}-\d{2}-\d{2}", compact)
        if any_iso_date:
            metadata["travel_date"] = any_iso_date.group(0)

    if "滴滴出行-行程单" in compact or "DIDI TRAVEL - TRIP TABLE" in compact:
        metadata["content_type"] = "打车行程单"
    elif "didi" in compact.lower() or "滴滴" in compact:
        metadata["content_type"] = "打车发票"
    elif "铁路电子客票" in compact:
        metadata["content_type"] = "高铁发票"
    elif "餐饮服务" in compact or "餐费" in compact:
        metadata["content_type"] = "餐饮"
    elif "汽油" in compact or "成品油" in compact:
        metadata["content_type"] = "油票"
    elif "住宿费" in compact:
        metadata["content_type"] = "住宿"
    elif "通行费" in compact:
        metadata["content_type"] = "过路费"

    return metadata


def read_ledger(path: Path) -> list[dict[str, Any]]:
    formulas = load_workbook(path, data_only=False)
    values = load_workbook(path, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet_name in formulas.sheetnames:
        ws = formulas[sheet_name]
        value_ws = values[sheet_name]
        headers = [normalize_header(ws.cell(1, col).value, col) for col in range(1, ws.max_column + 1)]
        for row_idx in range(2, ws.max_row + 1):
            raw_values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
            cached_values = [value_ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
            if not any(value not in (None, "") for value in raw_values):
                continue
            item: dict[str, Any] = {"source_sheet": sheet_name, "source_row": row_idx}
            for header, raw_value, cached_value in zip(headers, raw_values, cached_values):
                item[header] = cell_value(raw_value)
                item[f"{header}_value"] = cell_value(cached_value)
            rows.append(item)
    return rows


def collect_files(root: Path) -> list[FileInfo]:
    excluded = {
        "OA报销月度工作台.xlsx",
        "OA报销协作说明.md",
        "GZ＆HF&SZ报销指引  1.pdf",
        "财务-员工报销-Hester Zhao-2025-12-23.pdf",
    }
    files: list[FileInfo] = []
    for file_path in root.rglob("*"):
        if not file_path.is_file() or file_path.name == ".DS_Store":
            continue
        if file_path.name.startswith("~$"):
            continue
        rel = file_path.relative_to(root).as_posix()
        if rel.startswith("工具/") or file_path.name.startswith(("OA月度工作台_", "OA草稿清单_")) or rel in excluded:
            continue
        files.append(
            FileInfo(
                path=rel,
                ext=file_path.suffix.lower(),
                size=file_path.stat().st_size,
                categories=classify_file(rel),
                metadata=parse_pdf_metadata(file_path),
            )
        )
    return files


def normalize_date(value: Any, year: str) -> str:
    text = text_value(value)
    if not text:
        return ""
    match = re.match(r"^(\d{1,2})[.月/-](\d{1,2})", text)
    if match:
        return f"{year}-{match.group(1).zfill(2)}-{match.group(2).zfill(2)}"
    return text


def chinese_date_to_iso(value: Any) -> str:
    match = re.search(r"(20\d{2})年(\d{1,2})月(\d{1,2})日", text_value(value))
    if not match:
        return ""
    return f"{match.group(1)}-{match.group(2).zfill(2)}-{match.group(3).zfill(2)}"


def row_month(row: dict[str, Any], year: str) -> str:
    date = normalize_date(row.get("日期"), year)
    match = re.match(r"^(\d{4})-(\d{2})-", date)
    return "".join(match.groups()) if match else ""


def file_month(file_info: FileInfo, month: str) -> str:
    year = month[:4]
    month_no = int(month[4:6])
    month2 = month[4:6]
    short_year = year[2:]
    travel_date = text_value(file_info.metadata.get("travel_date"))
    if travel_date:
        return travel_date[:7].replace("-", "")
    invoice_date = chinese_date_to_iso(file_info.metadata.get("invoice_date"))
    if invoice_date:
        return invoice_date[:7].replace("-", "")
    if f"{year}.{month_no}" in file_info.path or f"{year}.{month2}" in file_info.path:
        return month
    if month in file_info.path or f"{year}.{month2}" in file_info.path or f"{year}-{month2}" in file_info.path:
        return month
    basename = Path(file_info.path).name
    if re.match(rf"^{short_year}{month2}\d{{2}}[_-]", basename):
        return month
    return ""


def monthly_files(files: list[FileInfo], month: str) -> list[FileInfo]:
    direct = [file_info for file_info in files if file_month(file_info, month) == month]
    parent_dirs = {"/".join(file_info.path.split("/")[:-1]) for file_info in direct}
    peers = [file_info for file_info in files if "/".join(file_info.path.split("/")[:-1]) in parent_dirs]
    by_path = {file_info.path: file_info for file_info in direct + peers}
    return sorted(by_path.values(), key=lambda item: item.path)


def categories_for_expense(key: str) -> list[str]:
    return {
        "油费": ["油票"],
        "过路费": ["过路费"],
        "餐饮费": ["餐费/交际"],
        "礼品费": ["餐费/交际"],
        "打车": ["交通", "打车发票", "打车行程单"],
        "停车": ["停车费"],
        "住宿": ["住宿"],
    }.get(key, [])


def date_patterns(value: Any, year: str) -> list[str]:
    normalized = normalize_date(value, year)
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", normalized)
    if not match:
        return []
    y, m, d = match.groups()
    return [
        f"{y}{m}{d}",
        f"{y[2:]}{m}{d}",
        f"{y}.{m}.{d}",
        f"{y}-{m}-{d}",
        f"{int(m)}.{int(d)}",
        f"{m}.{d}",
        f"{y}年{int(m)}月{int(d)}日",
        f"{y}年{m}月{d}日",
    ]


def merchant_from_filename(file_info: FileInfo) -> str:
    stem = Path(file_info.path).stem
    match = re.match(r"^\d{6}_\d+(?:\.\d{1,2})?_(.+)$", stem)
    return match.group(1) if match else ""


def suggest_files(files: list[FileInfo], key: str, row: dict[str, Any], amount: float | None, year: str) -> list[FileInfo]:
    wanted = categories_for_expense(key)
    patterns = date_patterns(row.get("日期"), year)
    amount_variants = [f"{amount:.2f}", str(amount)] if amount else []
    scored: list[tuple[int, FileInfo]] = []
    for file_info in files:
        score = 0
        if any(category in wanted for category in file_info.categories):
            score += 4
        content_type = file_info.metadata.get("content_type")
        if content_type and content_type in wanted:
            score += 3
        if any(pattern in file_info.path for pattern in patterns):
            score += 3
        if amount and float(file_info.metadata.get("amount") or 0) == amount:
            score += 3
        if any(variant in file_info.path for variant in amount_variants):
            score += 2
        if score >= 4:
            scored.append((score, file_info))
    return [item for _, item in sorted(scored, key=lambda pair: (-pair[0], pair[1].path))[:6]]


def make_oa_rows(month_rows: list[dict[str, Any]], files: list[FileInfo], year: str) -> tuple[list[list[Any]], list[list[Any]]]:
    oa_rows: list[list[Any]] = []
    unresolved: list[list[Any]] = []
    for row in month_rows:
        generated = 0
        for exp in EXPENSE_COLUMNS:
            raw = row.get(exp["key"])
            cached = row.get(f"{exp['key']}_value")
            amount = number_value(cached if cached not in ("", None) else raw)
            has_non_numeric = text_value(raw) and amount is None
            km = number_value(row.get("公里数_value") or row.get("公里数"))
            if amount and amount > 0:
                candidates = suggest_files(files, exp["key"], row, amount, year)
                candidate_text = "\n".join(candidate.path for candidate in candidates)
                missing: list[str] = []
                if exp["key"] == "油费" and not candidate_text:
                    missing.append("高德截图/油票待匹配")
                if exp["key"] == "过路费" and not candidate_text:
                    missing.append("ETC汇总单/通行费发票待补")
                if exp["key"] == "餐饮费":
                    missing.append("客户公司、客户姓名、同事姓名、午/晚餐待补")
                    if not candidate_text:
                        missing.append("餐饮发票待匹配")
                if exp["key"] == "打车" and not candidate_text:
                    missing.append("打车行程单/发票待匹配")
                if exp["key"] == "停车" and not candidate_text:
                    missing.append("停车票/付款记录待匹配")
                if exp["key"] == "住宿" and not candidate_text:
                    missing.append("酒店水单/发票待匹配")
                customer = text_value(row.get("客户"))
                description = "；".join(
                    part
                    for part in [
                        customer,
                        f"私车公用 {km}km" if exp["key"] == "油费" and km else "",
                        "客户/同事名单请补全" if exp["key"] == "餐饮费" else "",
                    ]
                    if part
                )
                oa_rows.append(
                    [
                        "待补材料/确认" if missing else "可录入草稿",
                        f"{row['source_sheet']} 行{row['source_row']}",
                        normalize_date(row.get("日期"), year),
                        customer,
                        exp["oa_category"],
                        exp["detail_type"],
                        "",
                        "",
                        km if exp["key"] == "油费" else "",
                        amount,
                        description,
                        "",
                        exp["attachments"],
                        candidate_text,
                        "；".join(missing),
                        exp["risk"],
                    ]
                )
                generated += 1
            elif has_non_numeric and exp["key"] != "住宿":
                unresolved.append(
                    [
                        "高",
                        f"{row['source_sheet']} 行{row['source_row']}",
                        normalize_date(row.get("日期"), year),
                        text_value(row.get("客户")),
                        exp["key"],
                        text_value(raw),
                        "清单中有内容但不是金额，需补金额或确认不报销。",
                    ]
                )
        if not generated and (text_value(row.get("客户")) or text_value(row.get("日期"))):
            unresolved.append(
                [
                    "中",
                    f"{row['source_sheet']} 行{row['source_row']}",
                    normalize_date(row.get("日期"), year),
                    text_value(row.get("客户")),
                    "未生成OA行",
                    "",
                    "清单无可用金额；如需报销请补金额和费用类型。",
                ]
            )
    return oa_rows, unresolved


def has_matching_transport_row(oa_rows: list[list[Any]], date: str, amount: float, detail_type: str) -> bool:
    for row in oa_rows:
        same_date = row[2] == date
        same_amount = abs(float(row[9] or 0) - amount) < 0.01
        same_detail = detail_type in text_value(row[5]) or text_value(row[5]) in detail_type
        if same_date and same_amount and same_detail:
            return True
    return False


def add_orphan_transport_rows(oa_rows: list[list[Any]], files: list[FileInfo]) -> None:
    didi_groups: dict[tuple[str, float], list[FileInfo]] = defaultdict(list)
    rail_files: list[FileInfo] = []

    for file_info in files:
        content_type = text_value(file_info.metadata.get("content_type"))
        travel_date = text_value(file_info.metadata.get("travel_date"))
        amount = number_value(file_info.metadata.get("amount"))
        if not travel_date or not amount:
            continue
        if content_type in {"打车发票", "打车行程单"}:
            didi_groups[(travel_date, amount)].append(file_info)
        elif content_type == "高铁发票":
            rail_files.append(file_info)

    for (travel_date, amount), group in sorted(didi_groups.items()):
        if has_matching_transport_row(oa_rows, travel_date, amount, "打车软件"):
            continue
        types = {text_value(item.metadata.get("content_type")) for item in group}
        missing = []
        if "打车发票" not in types:
            missing.append("滴滴电子发票待补")
        if "打车行程单" not in types:
            missing.append("滴滴出行行程报销单待补")
        missing.append("清单无对应打车金额/客户事项待确认")
        oa_rows.append(
            [
                "待补材料/确认",
                "文件自动识别",
                travel_date,
                "客户/事项待补",
                "交通费（含跨地区汽车费）",
                "打车软件",
                "",
                "",
                "",
                amount,
                "滴滴出行；客户/事项待补",
                "",
                "滴滴电子发票；滴滴出行行程报销单",
                "\n".join(item.path for item in group),
                "；".join(missing),
                "滴滴需发票和行程单成对上传；清单无对应行时先确认是否补入OA。",
            ]
        )

    for file_info in sorted(rail_files, key=lambda item: item.path):
        travel_date = text_value(file_info.metadata.get("travel_date"))
        amount = number_value(file_info.metadata.get("amount"))
        if not travel_date or not amount:
            continue
        if has_matching_transport_row(oa_rows, travel_date, amount, "火车"):
            continue
        oa_rows.append(
            [
                "待补材料/确认",
                "文件自动识别",
                travel_date,
                "客户/事项待补",
                "交通费（含跨地区汽车费）",
                "火车",
                "",
                "",
                "",
                amount,
                "高铁/铁路；客户/事项待补",
                "",
                "铁路电子客票/高铁发票",
                file_info.path,
                "清单无对应高铁金额/客户事项待确认",
                "高铁按乘车日期归月；OA附件上传铁路电子客票/高铁发票。",
            ]
        )


def style_range(ws, min_row: int, max_row: int, min_col: int, max_col: int, header_fill: str = "D9EAF7") -> None:
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[min_row][min_col - 1 : max_col]:
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_title(ws, title: str, subtitle: str, columns: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    ws.cell(1, 1, title)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor="1F4E79")
    ws.cell(1, 1).font = Font(color="FFFFFF", bold=True, size=16)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).fill = PatternFill("solid", fgColor="EAF3F8")
    ws.cell(2, 1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28


def write_table(ws, start_row: int, rows: list[list[Any]], widths: list[int], header_fill: str = "D9EAF7") -> None:
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)
    if rows:
        style_range(ws, start_row, start_row + len(rows) - 1, 1, len(rows[0]), header_fill)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_rules(ws) -> None:
    add_title(ws, "报销规则速查", "用于月度 OA 填报前自检。", 6)
    rows = [
        ["费用类型", "OA填写口径", "必备材料", "金额/标准", "常见退回风险", "备注"],
        ["私车公用-油费", "交通费；按日期/客户行程填写公里数和金额", "高德历史导航截图；成品油发票", "1km = 1.4元；油票金额需覆盖报销额", "无两地信息/无公里数/油票不足/日期不接近", "销售自驾出差通常需拜访两个或以上客户。"],
        ["过路费", "交通费；私车公用相关费用", "ETC电子票据汇总单；通行费电子发票", "按实际发生日期填写", "无汇总单；日期与行程不一致", "不同日期路桥费分开填写。"],
        ["打车软件", "交通费；打车软件", "行程单；电子发票", "通常 Max RMB 50/程，特殊情况写说明", "只有发票无行程单；起终点不清", "金额为“？”的先不要录入。"],
        ["交际应酬费-餐费", "交际应酬费；注明客户公司、人员、午/晚餐", "餐饮发票；必要时付款记录", "午餐 Max RMB 80/人；晚餐 Max RMB 200/人", "缺客户名单/业务背景；后开票无付款记录", "客户来访或无拜访记录时要写发生背景。"],
        ["住宿费", "住宿费；填写入住/退房日期、天数、地点", "酒店水单+发票；无水单则订单截图/批准", "AL5及others通常 Max RMB 400/day；AL4 Manager Max RMB 700/day", "无水单；超标无原因；日期不一致", "超额原因写在 OA 对应字段。"],
        ["停车费", "交通费；停车费单独列", "停车票/电子发票；必要时付款记录", "按实际发生", "混入油费；缺地点/日期", "与对应拜访日期放在同一报销批次。"],
    ]
    write_table(ws, 4, rows, [18, 28, 30, 25, 35, 28])


def build_workbook(root: Path, ledger_path: Path, month: str, output: Path) -> dict[str, Any]:
    year = month[:4]
    rows = read_ledger(ledger_path)
    all_files = collect_files(root)
    files = monthly_files(all_files, month)
    month_rows = [row for row in rows if row_month(row, year) == month]
    oa_rows, unresolved = make_oa_rows(month_rows, files, year)
    add_orphan_transport_rows(oa_rows, files)

    wb = Workbook()
    wb.remove(wb.active)

    summary = defaultdict(lambda: {"count": 0, "amount": 0.0})
    for row in oa_rows:
        summary[row[4]]["count"] += 1
        summary[row[4]]["amount"] += float(row[9] or 0)

    ws = wb.create_sheet("总览")
    add_title(ws, f"{month} OA月度工作台", "先处理查漏清单，再按 OA填报清单 录入并保存草稿；最终提交由用户确认。", 8)
    summary_rows = [["OA费用大类", "草稿行数", "草稿金额", "说明"]]
    for category, value in summary.items():
        summary_rows.append([category, value["count"], round(value["amount"], 2), "来自当前报销清单的月度数据"])
    summary_rows.append(["月度文件", len(files), "", "包含归入月份文件夹以及文件名/发票日期匹配目标月份的候选文件"])
    summary_rows.append(["待补/确认", sum("待补" in str(row[0]) for row in oa_rows) + len(unresolved), "", "优先看 查漏清单 和 OA填报清单 的黄色列"])
    write_table(ws, 4, summary_rows, [28, 12, 14, 60], "CFE8D5")

    ws = wb.create_sheet("OA填报清单")
    add_title(ws, f"{month} OA填报清单", "黄色列是提交前需要补材料或确认的地方。", 16)
    oa_headers = ["状态", "来源", "日期", "客户/事项", "OA费用大类", "OA明细类型", "起始地点", "终止地点", "公里数", "金额", "描述", "客户/同事名单", "应上传附件", "候选文件", "缺失材料", "退回风险"]
    write_table(ws, 4, [oa_headers] + oa_rows, [16, 12, 12, 20, 24, 24, 14, 14, 10, 12, 36, 30, 36, 48, 36, 36])
    for row_idx in range(5, 5 + len(oa_rows)):
        ws.cell(row_idx, 10).number_format = "¥#,##0.00"
        if "待补" in text_value(ws.cell(row_idx, 1).value):
            ws.cell(row_idx, 1).fill = PatternFill("solid", fgColor="FDE68A")
        ws.cell(row_idx, 15).fill = PatternFill("solid", fgColor="FFF7D6")
        ws.cell(row_idx, 16).fill = PatternFill("solid", fgColor="FFF7D6")

    ws = wb.create_sheet("原始清单记录")
    add_title(ws, f"{month} 原始清单记录", "从报销清单读取的月度行，便于回看来源。", 14)
    source_headers = ["来源", "日期", "客户", "公里数", "油费", "打车", "过路费", "餐饮费", "礼品费", "停车", "住宿", "补贴", "总费用", "备注"]
    source_rows = [
        [
            f"{row['source_sheet']} 行{row['source_row']}",
            normalize_date(row.get("日期"), year),
            text_value(row.get("客户")),
            text_value(row.get("公里数_value") or row.get("公里数")),
            text_value(row.get("油费_value") or row.get("油费")),
            text_value(row.get("打车_value") or row.get("打车")),
            text_value(row.get("过路费_value") or row.get("过路费")),
            text_value(row.get("餐饮费_value") or row.get("餐饮费")),
            text_value(row.get("礼品费_value") or row.get("礼品费")),
            text_value(row.get("停车_value") or row.get("停车")),
            text_value(row.get("住宿_value") or row.get("住宿")),
            text_value(row.get("补贴_value") or row.get("补贴")),
            text_value(row.get("总费用_value") or row.get("总费用")),
            text_value(row.get("备注_value") or row.get("备注")),
        ]
        for row in month_rows
    ]
    write_table(ws, 4, [source_headers] + source_rows, [12, 12, 22, 12, 12, 12, 12, 12, 12, 10, 14, 10, 12, 24], "DDEBF7")

    ws = wb.create_sheet("发票索引")
    add_title(ws, f"{month} 发票索引", "按月份、文件名和PDF可读信息生成；继续归档发票后可重新运行更新。", 9)
    file_headers = ["路径", "识别类型", "发票日期", "金额", "销售方/商户", "购买方", "发票号", "大小KB", "备注"]
    file_rows = [
        [
            file_info.path,
            file_info.metadata.get("content_type") or "；".join(file_info.categories) or "未分类",
            file_info.metadata.get("invoice_date", ""),
            file_info.metadata.get("amount", ""),
            merchant_from_filename(file_info) or file_info.metadata.get("seller", ""),
            file_info.metadata.get("buyer", ""),
            file_info.metadata.get("invoice_no", ""),
            round(file_info.size / 1024),
            "",
        ]
        for file_info in files
    ]
    write_table(ws, 4, [file_headers] + file_rows, [60, 18, 15, 12, 32, 32, 20, 10, 18], "DDEBF7")

    ws = wb.create_sheet("查漏清单")
    add_title(ws, f"{month} 查漏清单", "优先处理这些项目，能明显降低 OA 被退回概率。", 8)
    gap_headers = ["优先级", "来源", "日期", "客户/事项", "问题类型", "当前内容", "处理动作/说明", "状态"]
    gap_rows = [["高", row[1], row[2], row[3], row[5], row[9], row[14], "待处理"] for row in oa_rows if row[14]]
    gap_rows.extend([row + ["待处理"] for row in unresolved])
    write_table(ws, 4, [gap_headers] + gap_rows, [10, 12, 12, 20, 18, 14, 60, 12], "FDECC8")

    add_rules(wb.create_sheet("规则速查"))

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {
        "output": str(output),
        "month": month,
        "ledger_rows": len(month_rows),
        "oa_rows": len(oa_rows),
        "files": len(files),
        "gaps": len(gap_rows),
        "total": round(sum(float(row[9] or 0) for row in oa_rows), 2),
    }


def default_output(root: Path, month: str) -> Path:
    year = month[:4]
    month_number = int(month[4:6])
    return root / year / f"{year}.{month_number}" / f"OA月度工作台_{month}.xlsx"


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a monthly OA reimbursement workbook.")
    parser.add_argument("--root", default="/Users/hehe/Documents/报销", help="Reimbursement project folder.")
    parser.add_argument("--month", required=True, help="Target month in YYYYMM, e.g. 202605.")
    parser.add_argument("--ledger", default=None, help="Ledger workbook path. Defaults to <root>/<year>/报销清单.xlsx.")
    parser.add_argument("--output", default=None, help="Output workbook path.")
    parser.add_argument("--json", action="store_true", help="Print machine-readable JSON.")
    args = parser.parse_args()

    root = Path(args.root).expanduser().resolve()
    ledger = Path(args.ledger).expanduser().resolve() if args.ledger else root / args.month[:4] / "报销清单.xlsx"
    output = Path(args.output).expanduser().resolve() if args.output else default_output(root, args.month)
    result = build_workbook(root, ledger, args.month, output)

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(f"Created {result['output']}")
        print(f"OA rows: {result['oa_rows']}, total: {result['total']}, gaps: {result['gaps']}, files: {result['files']}")


if __name__ == "__main__":
    main()
